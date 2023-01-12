import openpyxl
import GameClass

def dataSifter():
    data = openpyxl.load_workbook("Miles_Mikolas_AdvancedPitchingStats.xlsx")
    sh = data.active

    #initializing some variables to use
    date = ""
    games = {}
    games["game1"] = GameClass.game()
    count = 0

    #appends elements to current game object lists
    def addElements(currRow):
        games["game{0}".format(count)].addPitches(sh.cell(row = currRow, column = 1).value)
        games["game{0}".format(count)].addEvents(sh.cell(row = currRow, column = 10).value)
        games["game{0}".format(count)].addOutcomes(sh.cell(row = currRow, column = 9).value)

    #for each row, if the date is not the same then create a new game object, append the data
    for i in range(2, sh.max_row+1):
        if sh.cell(row = i, column = 2).value != date:
            count += 1
            date = sh.cell(row = i, column = 2).value
            games["game{0}".format(count)] = GameClass.game()
        addElements(i)

    return games
